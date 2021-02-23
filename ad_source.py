#!/usr/bin/env python
# coding: utf-8

# In[65]:


import pandas as pd
import numpy as np
from openpyxl import load_workbook


# In[27]:


source = pd.read_csv("trafficSource_2018-06-26.csv")


# In[28]:


source.head(10)


# In[29]:


for i in range(27,31):
    path = "trafficSource_2018-06-"+ str(i) +".csv"
    source1 = pd.read_csv(path)
    source = pd.concat([source,source1], axis = 0, ignore_index = True)
source.drop


# In[30]:


for k in [7,8]:
    for i in range(1,10):
        path = "trafficSource_2018-0"+str(k)+"-0"+ str(i) +".csv"
        source1 = pd.read_csv(path)
        source = pd.concat([source,source1], axis = 0, ignore_index = True)
    for i in range(10,32):
        path = "trafficSource_2018-0"+str(k)+"-"+ str(i) +".csv"
        source1 = pd.read_csv(path)
        source = pd.concat([source,source1], axis = 0, ignore_index = True)


# In[31]:


for k in range(9,13):
    if k == 9:
        for i in range(1,10):
            path = "trafficSource_2018-0"+str(k)+"-0"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
        for i in range(10,31):
            path = "trafficSource_2018-0"+str(k)+"-"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True) 
    elif k%2 == 0:
        for i in range(1,10):
            path = "trafficSource_2018-"+str(k)+"-0"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
        for i in range(10,32):
            path = "trafficSource_2018-"+str(k)+"-"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True) 
    elif k%2 == 1:
        for i in range(1,10):
            path = "trafficSource_2018-"+str(k)+"-0"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True) 
        for i in range(10,31):
            path = "trafficSource_2018-"+str(k)+"-"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True) 


# In[52]:


source = source.drop(["mid","dc","did","ul","geoid","cm","cn"], axis = 1)


# In[53]:


data = source.dropna(subset=['uid'])
data1 = pd.concat([data], axis = 0, ignore_index = True)
data1


# In[54]:


source = pd.DataFrame(columns=['mid', 'dc', 'did', 'uid','ht','ul','geoid','bh','cs','cm','cn'])
for k in range(1,10):
    if k == 2:
        for i in range(1,10):
            path = "trafficSource_2019-0"+str(k)+"-0"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
        for i in range(10,29):
            path = "trafficSource_2019-0"+str(k)+"-"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
    elif (k%2 == 0) and (k != 8):
        for i in range(1,10):
            path = "trafficSource_2019-0"+str(k)+"-0"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
        for i in range(10,31):
            path = "trafficSource_2019-0"+str(k)+"-"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
    elif (k%2 == 1) and (k != 9):
        for i in range(1,10):
            path = "trafficSource_2019-0"+str(k)+"-0"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
        for i in range(10,32):
            path = "trafficSource_2019-0"+str(k)+"-"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
    elif k == 8:
        for i in range(1,10):
            path = "trafficSource_2019-0"+str(k)+"-0"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
        for i in range(10,32):
            path = "trafficSource_2019-0"+str(k)+"-"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
    elif k == 9:
        for i in range(1,10):
            path = "trafficSource_2019-0"+str(k)+"-0"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
        for i in range(10,31):
            path = "trafficSource_2019-0"+str(k)+"-"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)


# In[55]:


for k in range(10,13):
    if k == 11:
        for i in range(1,10):
            path = "trafficSource_2019-"+str(k)+"-0"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
        for i in range(10,31):
            path = "trafficSource_2019-"+str(k)+"-"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
    else:
        for i in range(1,10):
            path = "trafficSource_2019-"+str(k)+"-0"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)
        for i in range(10,31):
            path = "trafficSource_2019-"+str(k)+"-"+ str(i) +".csv"
            source1 = pd.read_csv(path)
            source = pd.concat([source,source1], axis = 0, ignore_index = True)


# In[56]:


source = source.drop(["mid","dc","did","ul","geoid","cm","cn"], axis = 1)
data = source.dropna(subset=['uid'])
data2 = pd.concat([data,data1], axis = 0, ignore_index = True)
data2


# In[62]:


source = pd.read_csv("trafficSource_2020-01-01.csv")


# In[63]:


for i in range(2,10):
    path = "trafficSource_2020-01-0"+ str(i) +".csv"
    source1 = pd.read_csv(path, encoding='ISO-8859-1')
    source = pd.concat([source,source1], axis = 0, ignore_index = True)
for i in range(10,32):
    path = "trafficSource_2020-01-"+ str(i) +".csv"
    source1 = pd.read_csv(path, encoding='ISO-8859-1')
    source = pd.concat([source,source1], axis = 0, ignore_index = True)

for i in range(1,10):
    path = "trafficSource_2020-02-0"+ str(i) +".csv"
    source1 = pd.read_csv(path)
    source = pd.concat([source,source1], axis = 0, ignore_index = True)
for i in range(10,30):
    path = "trafficSource_2020-02-"+ str(i) +".csv"
    source1 = pd.read_csv(path, encoding='ISO-8859-1')
    source = pd.concat([source,source1], axis = 0, ignore_index = True)

for i in range(1,10):
    path = "trafficSource_2020-03-0"+ str(i) +".csv"
    source1 = pd.read_csv(path, encoding='ISO-8859-1')
    source = pd.concat([source,source1], axis = 0, ignore_index = True)
for i in range(10,32):
    path = "trafficSource_2020-03-"+ str(i) +".csv"
    source1 = pd.read_csv(path, encoding='ISO-8859-1')
    source = pd.concat([source,source1], axis = 0, ignore_index = True)

for i in range(1,10):
    path = "trafficSource_2020-04-0"+ str(i) +".csv"
    source1 = pd.read_csv(path, encoding='ISO-8859-1')
    source = pd.concat([source,source1], axis = 0, ignore_index = True)
for i in range(10,31):
    path = "trafficSource_2020-04-"+ str(i) +".csv"
    source1 = pd.read_csv(path, encoding='ISO-8859-1')
    source = pd.concat([source,source1], axis = 0, ignore_index = True)


# In[64]:


source = source.drop(["mid","dc","did","ul","geoid","cm","cn"], axis = 1)
data = source.dropna(subset=['uid'])
data3 = pd.concat([data2,data], axis = 0, ignore_index = True)
data3


# In[66]:


result = load_workbook(filename = "Source.xlsx")
result_sheet = result["工作表1"]

for i in range(4217548):
    result_sheet.cell(row=i+2, column=1).value = data3["uid"][i]
    result_sheet.cell(row=i+2, column=2).value = data3["ht"][i]
    result_sheet.cell(row=i+2, column=3).value = data3["bh"][i]
    result_sheet.cell(row=i+2, column=4).value = data3["cs"][i]

result.save(filename = "Source.xlsx")

