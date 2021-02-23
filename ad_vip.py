#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pandas as pd
from openpyxl import load_workbook
import datetime


# In[6]:


data = pd.read_excel("Data.xlsx")


# In[7]:


data.head()


# In[9]:


set(data["cs"])


# In[10]:


set(data["MemberCardLevel"])


# In[11]:


set(data["bh"])


# In[14]:


g1 = (data["MemberCardLevel"] == 10)
g10 = data[g1]
g10.reset_index(drop=True, inplace=True)
g10


# In[15]:


g1 = (data["MemberCardLevel"] == 20)
g20 = data[g1]
g20.reset_index(drop=True, inplace=True)
g20


# In[16]:


g1 = (data["MemberCardLevel"] == 30)
g30 = data[g1]
g30.reset_index(drop=True, inplace=True)
g30


# In[17]:


g1 = (data["MemberCardLevel"] == 40)
g40 = data[g1]
g40.reset_index(drop=True, inplace=True)
g40


# In[20]:


cs10 = {}
cs20 = {}
cs30 = {}
cs40 = {}


# In[19]:


ID_list = []
cs_dict = {}
cs_kind = g10.drop_duplicates(subset='cs', keep='first', inplace=False)
cs_kind.reset_index(drop=True, inplace=True)
for i in range(len(cs_kind["cs"])):
    cs_dict[cs_kind["cs"][i]] = 0
cs_dict


# In[21]:


for i in range(len(g10["MemberID"])):
    if g10["MemberID"][i] not in ID_list:
        ID_list.append(g10["MemberID"][i])
        cs_dict[g10["cs"][i]] += 2
    cs_dict[g10["cs"][i]] += 1
cs10 = cs_dict


# In[22]:


ID_list = []
cs_dict = {}
cs_kind = g20.drop_duplicates(subset='cs', keep='first', inplace=False)
cs_kind.reset_index(drop=True, inplace=True)
for i in range(len(cs_kind["cs"])):
    cs_dict[cs_kind["cs"][i]] = 0
for i in range(len(g20["MemberID"])):
    if g20["MemberID"][i] not in ID_list:
        ID_list.append(g20["MemberID"][i])
        cs_dict[g20["cs"][i]] += 2
    cs_dict[g20["cs"][i]] += 1
cs20 = cs_dict


# In[23]:


ID_list = []
cs_dict = {}
cs_kind = g30.drop_duplicates(subset='cs', keep='first', inplace=False)
cs_kind.reset_index(drop=True, inplace=True)
for i in range(len(cs_kind["cs"])):
    cs_dict[cs_kind["cs"][i]] = 0
for i in range(len(g30["MemberID"])):
    if g30["MemberID"][i] not in ID_list:
        ID_list.append(g30["MemberID"][i])
        cs_dict[g30["cs"][i]] += 2
    cs_dict[g30["cs"][i]] += 1
cs30 = cs_dict


# In[24]:


ID_list = []
cs_dict = {}
cs_kind = g40.drop_duplicates(subset='cs', keep='first', inplace=False)
cs_kind.reset_index(drop=True, inplace=True)
for i in range(len(cs_kind["cs"])):
    cs_dict[cs_kind["cs"][i]] = 0
for i in range(len(g40["MemberID"])):
    if g40["MemberID"][i] not in ID_list:
        ID_list.append(g40["MemberID"][i])
        cs_dict[g40["cs"][i]] += 2
    cs_dict[g40["cs"][i]] += 1
cs40 = cs_dict


# In[26]:


cs40


# In[33]:


Cs10 = pd.DataFrame(list(cs10.items()), columns=['cs_source', 'score'])
Cs20 = pd.DataFrame(list(cs20.items()), columns=['cs_source', 'score'])
Cs30 = pd.DataFrame(list(cs30.items()), columns=['cs_source', 'score'])
Cs40 = pd.DataFrame(list(cs40.items()), columns=['cs_source', 'score'])


# In[35]:


result = load_workbook(filename = "Group_Source.xlsx")


# In[36]:


result = load_workbook(filename = "Group_Source.xlsx")

result_sheet = result["10"]
for i in range(len(Cs10["cs_source"])):
    result_sheet.cell(row=i+2, column=1).value = Cs10["cs_source"][i]
    result_sheet.cell(row=i+2, column=2).value = Cs10["score"][i]

result_sheet = result["20"]
for i in range(len(Cs20["cs_source"])):
    result_sheet.cell(row=i+2, column=1).value = Cs20["cs_source"][i]
    result_sheet.cell(row=i+2, column=2).value = Cs20["score"][i]

result_sheet = result["30"]
for i in range(len(Cs30["cs_source"])):
    result_sheet.cell(row=i+2, column=1).value = Cs30["cs_source"][i]
    result_sheet.cell(row=i+2, column=2).value = Cs30["score"][i]

result_sheet = result["40"]
for i in range(len(Cs40["cs_source"])):
    result_sheet.cell(row=i+2, column=1).value = Cs40["cs_source"][i]
    result_sheet.cell(row=i+2, column=2).value = Cs40["score"][i]
    
result.save(filename = "Group_Source.xlsx")


# In[ ]:




