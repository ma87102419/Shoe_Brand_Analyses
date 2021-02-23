#!/usr/bin/env python
# coding: utf-8

# In[20]:


import pandas as pd
from openpyxl import load_workbook
import datetime


# In[14]:


source = pd.read_csv("Source.csv")
member = pd.read_csv("MemberData.csv")


# In[15]:


member = member.drop(["RegisterSourceTypeDef","RegisterDateTime","Gender","IsAppInstalled","IsEnableEmail","IsEnablePushNotification","IsEnableShortMessage","eland_uuid"], axis = 1)


# In[16]:


member.head(10)


# In[13]:


member[member["MemberID"]=="%2B%2B3XVjBiQPLoAXWc7gK%2BTdvTH4J4AmgkeFlB3mmN7YY%3D"][["Birthday","MemberID"]]


# In[18]:


data = pd.merge(source,member, on=['MemberID'],how='inner')
data


# In[29]:


data = data.drop_duplicates(subset=None, keep='first', inplace=False)
data


# In[30]:


data = pd.concat([data], axis = 0, ignore_index = True)
data


# In[53]:


data.iloc[16278]


# In[54]:


for i in range(16279,52232):
    day = data["Birthday"][i]
    age = 2020 - int(day[:4])
    data["age"][i] = age
    ht = data["ht"][i]
    active_time = datetime.datetime.fromtimestamp(int(ht)/1000)
    data["ht"][i] = active_time
    print(i)


# In[55]:


data


# In[98]:


result = load_workbook(filename = "Data.xlsx")

result_sheet = result["工作表1"]
for i in range(len(data["MemberID"])):
    result_sheet.cell(row=i+2, column=1).value = data["MemberID"][i]
    result_sheet.cell(row=i+2, column=2).value = data["ht"][i]
    result_sheet.cell(row=i+2, column=3).value = data["bh"][i]
    result_sheet.cell(row=i+2, column=4).value = data["cs"][i]
    result_sheet.cell(row=i+2, column=5).value = data["age"][i]
    result_sheet.cell(row=i+2, column=6).value = data["MemberCardLevel"][i]

result.save(filename = "Data.xlsx")


# In[61]:


g1 = (10 <= data["age"])
g2 = data[g1]
g3 = (g2["age"] < 20)
group10_20 = g2[g3]
group10_20.reset_index(drop=True, inplace=True)
group10_20


# In[63]:


g1 = (20 <= data["age"])
g2 = data[g1]
g3 = (g2["age"] < 30)
group20_30 = g2[g3]
group20_30.reset_index(drop=True, inplace=True)
group20_30


# In[65]:


g1 = (30 <= data["age"])
g2 = data[g1]
g3 = (g2["age"] < 40)
group30_40 = g2[g3]
group30_40.reset_index(drop=True, inplace=True)
group30_40


# In[66]:


g1 = (40 <= data["age"])
group_40 = data[g1]
group_40.reset_index(drop=True, inplace=True)
group_40


# In[87]:


cs10_20 = {}
cs20_30 = {}
cs30_40 = {}
cs_40 = {}


# In[84]:


ID_list = []
cs_dict = {}
cs_kind = group10_20.drop_duplicates(subset='cs', keep='first', inplace=False)
cs_kind.reset_index(drop=True, inplace=True)
for i in range(len(cs_kind["cs"])):
    cs_dict[cs_kind["cs"][i]] = 0


# In[88]:


for i in range(len(group10_20["MemberID"])):
    if group10_20["MemberID"][i] not in ID_list:
        ID_list.append(group10_20["MemberID"][i])
        cs_dict[group10_20["cs"][i]] += 2
    cs_dict[group10_20["cs"][i]] += 1
cs10_20 = cs_dict


# In[90]:


ID_list = []
cs_dict = {}
cs_kind = group20_30.drop_duplicates(subset='cs', keep='first', inplace=False)
cs_kind.reset_index(drop=True, inplace=True)
for i in range(len(cs_kind["cs"])):
    cs_dict[cs_kind["cs"][i]] = 0
for i in range(len(group20_30["MemberID"])):
    if group20_30["MemberID"][i] not in ID_list:
        ID_list.append(group20_30["MemberID"][i])
        cs_dict[group20_30["cs"][i]] += 2
    cs_dict[group20_30["cs"][i]] += 1
cs20_30 = cs_dict


# In[92]:


ID_list = []
cs_dict = {}
cs_kind = group30_40.drop_duplicates(subset='cs', keep='first', inplace=False)
cs_kind.reset_index(drop=True, inplace=True)
for i in range(len(cs_kind["cs"])):
    cs_dict[cs_kind["cs"][i]] = 0
for i in range(len(group30_40["MemberID"])):
    if group30_40["MemberID"][i] not in ID_list:
        ID_list.append(group30_40["MemberID"][i])
        cs_dict[group30_40["cs"][i]] += 2
    cs_dict[group30_40["cs"][i]] += 1
cs30_40 = cs_dict


# In[94]:


ID_list = []
cs_dict = {}
cs_kind = group_40.drop_duplicates(subset='cs', keep='first', inplace=False)
cs_kind.reset_index(drop=True, inplace=True)
for i in range(len(cs_kind["cs"])):
    cs_dict[cs_kind["cs"][i]] = 0
for i in range(len(group_40["MemberID"])):
    if group_40["MemberID"][i] not in ID_list:
        ID_list.append(group_40["MemberID"][i])
        cs_dict[group_40["cs"][i]] += 2
    cs_dict[group_40["cs"][i]] += 1
cs_40 = cs_dict


# In[95]:


cs_40


# In[96]:


Cs10_20 = pd.DataFrame(list(cs10_20.items()), columns=['cs_source', 'score'])
Cs20_30 = pd.DataFrame(list(cs20_30.items()), columns=['cs_source', 'score'])
Cs30_40 = pd.DataFrame(list(cs30_40.items()), columns=['cs_source', 'score'])
Cs_40 = pd.DataFrame(list(cs_40.items()), columns=['cs_source', 'score'])


# In[97]:


result = load_workbook(filename = "Group_Source.xlsx")

result_sheet = result["10-20"]
for i in range(len(Cs10_20["cs_source"])):
    result_sheet.cell(row=i+2, column=1).value = Cs10_20["cs_source"][i]
    result_sheet.cell(row=i+2, column=2).value = Cs10_20["score"][i]

result_sheet = result["20-30"]
for i in range(len(Cs20_30["cs_source"])):
    result_sheet.cell(row=i+2, column=1).value = Cs20_30["cs_source"][i]
    result_sheet.cell(row=i+2, column=2).value = Cs20_30["score"][i]

result_sheet = result["30-40"]
for i in range(len(Cs30_40["cs_source"])):
    result_sheet.cell(row=i+2, column=1).value = Cs30_40["cs_source"][i]
    result_sheet.cell(row=i+2, column=2).value = Cs30_40["score"][i]

result_sheet = result["40"]
for i in range(len(Cs_40["cs_source"])):
    result_sheet.cell(row=i+2, column=1).value = Cs_40["cs_source"][i]
    result_sheet.cell(row=i+2, column=2).value = Cs_40["score"][i]
    
result.save(filename = "Group_Source.xlsx")

